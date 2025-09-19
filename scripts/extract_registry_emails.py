#!/usr/bin/env python3
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    sys.stderr.write("openpyxl is required to extract registry emails\n")
    sys.exit(2)


def normalize_email(value):
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    return value


def normalize_name(value):
    if not value:
        return None
    value = str(value).strip()
    return value if value else None


def extract(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: idx for idx, name in enumerate(header_row) if name}

    name_idx = headers.get("Entity Name")
    email_idx = headers.get("Contact Email")

    if name_idx is None or email_idx is None:
        raise RuntimeError("Expected columns 'Entity Name' and 'Contact Email' in registry workbook")

    records = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = normalize_name(row[name_idx] if name_idx < len(row) else None)
        email = normalize_email(row[email_idx] if email_idx < len(row) else None)
        if not name or not email:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        records.append({
            "name": name,
            "email": email,
        })

    return records


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("Usage: extract_registry_emails.py <path-to-xlsx>\n")
        sys.exit(1)

    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        sys.stderr.write(f"Registry workbook not found: {path}\n")
        sys.exit(1)

    records = extract(path)
    json.dump(records, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
